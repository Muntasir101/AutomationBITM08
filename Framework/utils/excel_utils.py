import openpyxl


def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_row


"""
row_count = get_row_count("E:\\Offline_Batch_08\\Projects\\AutomationBITM08\\Framework\\data\\test_data.xlsx", "Sheet1")
print(row_count)
"""


def get_column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_column


"""
column_count = get_column_count("E:\\Offline_Batch_08\\Projects\\AutomationBITM08\\Framework\\data\\test_data.xlsx",
                                "Sheet1")
print(column_count)
"""


def read_data(file, sheet, reading_row, reading_column):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading_row, column=reading_column).value


"""
data = read_data("E:\\Offline_Batch_08\\Projects\\AutomationBITM08\\Framework\\data\\test_data.xlsx",
                 "Sheet1", 2, 1)
print(data)
"""


def write_data(file, sheet, reading_row, reading_column, data):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    worksheet.cell(row=reading_row, column=reading_column).value = data
    workbook.save(file)


"""
write_data("E:\\Offline_Batch_08\\Projects\\AutomationBITM08\\Framework\\data\\test_data.xlsx",
           "Sheet1", 6, 1, "Hello Admin")
"""
